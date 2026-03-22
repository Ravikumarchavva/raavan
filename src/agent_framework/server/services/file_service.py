"""File service — save, list, retrieve and extract text from uploaded files.

Storage is delegated to the pluggable ``FileStore`` layer (local, S3, etc.).
Metadata (ownership, path, checksum) lives in the ``file_metadata`` table.

Text extraction supports:
  - CSV  → column summary + first 100 rows
  - JSON → pretty-printed
  - TXT / MD / source-code → raw UTF-8
  - PDF  → per-page text via pypdf  (optional dep)
  - XLSX → sheet rows via openpyxl  (optional dep)

Images are not converted to text here; the caller receives a separate
``image_url`` vision block that can be embedded in a multimodal message.
"""

from __future__ import annotations

import base64
import csv
import io
import json
import logging
import uuid
from typing import Optional

from sqlalchemy import select, update
from sqlalchemy.ext.asyncio import AsyncSession

from agent_framework.core.storage.base import FileRef, FileStore
from agent_framework.core.storage.tenant import FileScope, TenantContext
from agent_framework.server.models import FileMetadata

logger = logging.getLogger(__name__)

# Maximum characters injected per file into the LLM context window
_MAX_TEXT_CHARS = 50_000

_TEXT_MIMES = {
    "text/plain",
    "text/markdown",
    "text/html",
    "text/css",
    "text/javascript",
    "application/javascript",
    "application/x-python",
    "application/x-sh",
    "application/xml",
    "text/xml",
    "text/x-python",
}
_TEXT_EXTS = {
    "txt",
    "md",
    "py",
    "js",
    "ts",
    "tsx",
    "jsx",
    "html",
    "css",
    "sh",
    "yaml",
    "yml",
    "toml",
    "ini",
    "conf",
    "log",
    "r",
    "sql",
    "xml",
}

_XLSX_MIMES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
}

_IMAGE_MIMES = {
    "image/png",
    "image/jpeg",
    "image/gif",
    "image/webp",
    "image/svg+xml",
}


# ---------------------------------------------------------------------------
# CRUD helpers
# ---------------------------------------------------------------------------


async def save_file(
    db: AsyncSession,
    store: FileStore,
    *,
    thread_id: uuid.UUID,
    name: str,
    mime: str,
    content: bytes,
    user_id: uuid.UUID | None = None,
    org_id: str | None = None,
    scope: FileScope = FileScope.UPLOADS,
) -> FileMetadata:
    """Upload a file to the external store and record metadata in the DB."""
    tenant = TenantContext(
        org_id=org_id,
        user_id=str(user_id) if user_id else None,
        thread_id=str(thread_id),
    )
    object_key = tenant.key(name, scope)

    ref: FileRef = await store.put(
        object_key,
        content,
        content_type=mime,
    )

    meta = FileMetadata(
        thread_id=thread_id,
        user_id=user_id,
        org_id=org_id,
        scope=scope.value,
        object_key=ref.object_key,
        original_name=name,
        content_type=mime,
        size_bytes=ref.size_bytes,
        checksum_sha256=ref.checksum_sha256,
    )
    db.add(meta)
    await db.flush()
    return meta


async def list_files(
    db: AsyncSession,
    thread_id: uuid.UUID,
) -> list[FileMetadata]:
    """Return all non-deleted files attached to a thread, oldest first."""
    result = await db.execute(
        select(FileMetadata)
        .where(
            FileMetadata.thread_id == thread_id,
            FileMetadata.deleted_at.is_(None),
        )
        .order_by(FileMetadata.created_at)
    )
    return list(result.scalars().all())


async def get_file(
    db: AsyncSession,
    file_id: uuid.UUID,
    thread_id: uuid.UUID,
) -> Optional[FileMetadata]:
    """Get a single non-deleted file that belongs to the given thread."""
    result = await db.execute(
        select(FileMetadata).where(
            FileMetadata.id == file_id,
            FileMetadata.thread_id == thread_id,
            FileMetadata.deleted_at.is_(None),
        )
    )
    return result.scalar_one_or_none()


async def delete_file(
    db: AsyncSession,
    store: FileStore,
    file_id: uuid.UUID,
    thread_id: uuid.UUID,
) -> bool:
    """Soft-delete a file. Also removes the object from the store.

    Returns True if found and deleted.
    """
    from datetime import datetime, timezone

    meta = await get_file(db, file_id, thread_id)
    if not meta:
        return False

    # Remove from external store
    await store.delete(meta.object_key)

    # Soft-delete in DB
    meta.deleted_at = datetime.now(timezone.utc)
    await db.flush()
    return True


async def get_files_by_ids(
    db: AsyncSession,
    file_ids: list[uuid.UUID],
    thread_id: uuid.UUID,
) -> list[FileMetadata]:
    """Fetch a set of non-deleted files by ID, all belonging to the same thread."""
    if not file_ids:
        return []
    result = await db.execute(
        select(FileMetadata).where(
            FileMetadata.id.in_(file_ids),
            FileMetadata.thread_id == thread_id,
            FileMetadata.deleted_at.is_(None),
        )
    )
    return list(result.scalars().all())


async def get_file_content(
    store: FileStore,
    meta: FileMetadata,
) -> bytes:
    """Download the file bytes from the external store."""
    return await store.get(meta.object_key)


async def get_file_url(
    store: FileStore,
    meta: FileMetadata,
    *,
    expires_in: int = 3600,
) -> str:
    """Generate a pre-signed download URL for a file.

    For LocalFileStore this returns a ``file://`` URI.
    For S3 it returns a time-limited HTTPS pre-signed URL.
    """
    return await store.get_url(meta.object_key, expires_in=expires_in)


async def purge_thread_files(
    db: AsyncSession,
    store: FileStore,
    thread_id: uuid.UUID,
) -> int:
    """Remove all files (store + DB) for a given thread.

    Returns the number of objects deleted from the store.
    """
    from datetime import datetime, timezone

    files = await list_files(db, thread_id)
    if not files:
        return 0

    # Build prefix for bulk store deletion
    # (faster than deleting one by one)
    tenant = TenantContext(thread_id=str(thread_id))
    count = await store.delete_prefix(tenant.prefix())

    # Soft-delete all metadata rows
    now = datetime.now(timezone.utc)
    await db.execute(
        update(FileMetadata)
        .where(
            FileMetadata.thread_id == thread_id,
            FileMetadata.deleted_at.is_(None),
        )
        .values(deleted_at=now)
    )
    await db.flush()
    return count


# ---------------------------------------------------------------------------
# Text extraction
# ---------------------------------------------------------------------------


def extract_text_from_bytes(
    raw: bytes,
    *,
    name: str = "",
    mime: str = "",
) -> Optional[str]:
    """Extract a text representation from raw file bytes for LLM context.

    Returns the extracted string (≤ _MAX_TEXT_CHARS chars), or None when
    the file is binary-only (images / unknown blobs) and cannot be texified.

    This is a pure function — no DB or store access.
    """
    if not raw:
        return None

    mime = mime.lower()
    ext = name.rsplit(".", 1)[-1].lower() if "." in name else ""

    # ── Plain text / source code ──────────────────────────────────────────
    if mime in _TEXT_MIMES or ext in _TEXT_EXTS:
        try:
            return raw.decode("utf-8", errors="replace")[:_MAX_TEXT_CHARS]
        except Exception:
            return None

    # ── JSON ──────────────────────────────────────────────────────────────
    if mime == "application/json" or ext == "json":
        try:
            parsed = json.loads(raw.decode("utf-8"))
            return json.dumps(parsed, indent=2)[:_MAX_TEXT_CHARS]
        except Exception:
            return raw.decode("utf-8", errors="replace")[:_MAX_TEXT_CHARS]

    # ── CSV ───────────────────────────────────────────────────────────────
    if mime in ("text/csv", "application/csv") or ext == "csv":
        try:
            text = raw.decode("utf-8", errors="replace")
            reader = csv.reader(io.StringIO(text))
            rows = list(reader)
            if not rows:
                return "(empty CSV)"
            headers = rows[0]
            total_rows = len(rows) - 1
            preview = rows[1:101]
            lines: list[str] = [
                f"CSV file: {name}",
                f"Columns ({len(headers)}): {', '.join(headers)}",
                f"Total rows: {total_rows}",
                "",
                ",".join(headers),
            ]
            lines.extend(",".join(str(v) for v in r) for r in preview)
            if total_rows > 100:
                lines.append(f"... ({total_rows - 100} more rows not shown)")
            return "\n".join(lines)[:_MAX_TEXT_CHARS]
        except Exception as exc:
            logger.warning("CSV extraction failed for %s: %s", name, exc)
            return None

    # ── PDF ───────────────────────────────────────────────────────────────
    if mime == "application/pdf" or ext == "pdf":
        try:
            import pypdf  # noqa: PLC0415

            reader_obj = pypdf.PdfReader(io.BytesIO(raw))
            pages: list[str] = []
            for i, page in enumerate(reader_obj.pages):
                pages.append(f"--- Page {i + 1} ---\n{page.extract_text() or ''}")
            return "\n\n".join(pages)[:_MAX_TEXT_CHARS]
        except ImportError:
            logger.warning(
                "pypdf not installed; PDF text extraction unavailable for %s", name
            )
            return (
                f"(PDF file: {name} — text extraction unavailable, "
                "install pypdf to enable it. "
                "The file is available in the code interpreter at /data/{name})"
            )
        except Exception as exc:
            logger.warning("PDF extraction failed for %s: %s", name, exc)
            return None

    # ── XLSX ──────────────────────────────────────────────────────────────
    if mime in _XLSX_MIMES or ext in ("xlsx", "xls"):
        try:
            import openpyxl  # noqa: PLC0415

            wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            lines: list[str] = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                lines.append(f"=== Sheet: {sheet_name} ===")
                for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                    lines.append(",".join("" if v is None else str(v) for v in row))
                    if row_idx >= 199:
                        lines.append("... (truncated at 200 rows)")
                        break
            return "\n".join(lines)[:_MAX_TEXT_CHARS]
        except ImportError:
            logger.warning(
                "openpyxl not installed; XLSX text extraction unavailable for %s", name
            )
            return (
                f"(Excel file: {name} — text extraction unavailable, "
                "install openpyxl to enable it. "
                f"The file is available in the code interpreter at /data/{name})"
            )
        except Exception as exc:
            logger.warning("XLSX extraction failed for %s: %s", name, exc)
            return None

    # ── Images — no text; handled via vision blocks ───────────────────────
    if mime.startswith("image/"):
        return None

    # ── Unknown binary ────────────────────────────────────────────────────
    return None


async def extract_text(
    store: FileStore,
    meta: FileMetadata,
) -> Optional[str]:
    """Download file from store and extract text for LLM context."""
    raw = await store.get(meta.object_key)
    return extract_text_from_bytes(
        raw,
        name=meta.original_name,
        mime=meta.content_type,
    )


async def to_vision_image_block(
    store: FileStore,
    meta: FileMetadata,
) -> Optional[dict]:
    """Build an OpenAI-compatible vision content block from an image file.

    Returns None if the file is not an image.
    """
    mime = (meta.content_type or "image/png").lower()
    if not mime.startswith("image/"):
        return None

    raw = await store.get(meta.object_key)
    b64 = base64.b64encode(raw).decode()
    return {
        "type": "image_url",
        "image_url": {
            "url": f"data:{mime};base64,{b64}",
            "detail": "auto",
        },
    }
